# ==================== ✅ IMPORTAÇÕES ==================== #
# Bibliotecas padrão
import os
from decimal import Decimal, InvalidOperation

# Bibliotecas de terceiros
import openpyxl
import pdfkit
from openpyxl.utils import get_column_letter

# Django - Configurações e utilitários
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string

# Aplicação - Modelos e Formulários
from .forms import UsuarioForm, Usuario, EditarUsuarioForm
from .models import Agendamento, Pet, Servico, Tutor

# ==================== ✅ VIEWS DE AUTENTICAÇÃO ==================== #
@login_required
def home(request):
    """
    Exibe a página inicial do sistema para usuários autenticados.
    """
    return render(request, 'home.html')


User = get_user_model()


def registrar_usuario(request):
    """
    Cadastra um novo usuário no sistema.
    Permite cadastro livre apenas se ainda não existe nenhum usuário (primeiro admin).
    Depois disso, apenas usuários logados e com nível admin podem registrar.
    """

    # Verifica se já existem usuários no sistema
    primeiro_usuario = User.objects.count() == 0

    # Se não é o primeiro usuário, exige login e nível admin
    if not primeiro_usuario:
        if not request.user.is_authenticated:
            messages.error(request, "Você precisa estar logado para registrar novos usuários.")
            return redirect('login')

        if request.user.nivel != 'admin':
            messages.error(request, "Você não tem permissão para registrar novos usuários.")
            return redirect('home')

    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            usuario = form.save(commit=False)

            # Define o nível do novo usuário
            if primeiro_usuario:
                usuario.nivel = 'admin'  # Primeiro usuário é admin
            else:
                usuario.nivel = request.POST.get('nivel', 'usuario')  

            usuario.save()
            messages.success(request, 'Usuário registrado com sucesso!')

            return redirect('registrar')
        else:
            messages.error(request, 'Erro ao registrar usuário. Verifique os dados.')
    else:
        form = UsuarioForm()

    return render(request, 'registrar.html', {'form': form})


def login_view(request):
    """
    Realiza a autenticação do usuário e o login no sistema.

    - Processa o formulário de login (username e password).
    - Valida as credenciais do usuário.
    - Realiza o login em caso de sucesso e redireciona para a página inicial.
    - Exibe mensagem de erro em caso de falha.
    """
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            #messages.success(request, 'Login realizado com sucesso!')
            return redirect('home')
        else:
            messages.error(request, 'Usuário ou senha inválidos.')
    return render(request, 'login.html')


def logout_view(request):
    """
    Realiza o logout do usuário e o redireciona para a página de login.
    """
    logout(request)
    #messages.success(request, 'Logout realizado com sucesso!')
    return redirect('login')


# ==================== ✅ VIEWS DE USUÁRIOS ==================== #
@login_required
def listar_usuarios(request):
    '''
    Lista usuários cadastrados
    e faz pesquisa por nome
    '''
    query = request.GET.get('q', '').strip()
    if query:
        usuarios = Usuario.objects.filter(
            Q(username__icontains=query) | Q(first_name__icontains=query) | Q(last_name__icontains=query)
        )
    else:
        usuarios = Usuario.objects.all()
    return render(request, 'listar_usuarios.html', {'usuarios': usuarios})


@login_required
def editar_usuario(request, usuario_id):
    """
    Edita os dados de um usuário existente.
    """
    usuario = get_object_or_404(Usuario, id=usuario_id)
    if request.method == 'POST':
        form = EditarUsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Dados do usuário atualizados com sucesso!')
            return redirect('listar_usuarios')
        else:
            messages.error(request, 'Por favor, corrija os erros no formulário.')
    else:
        form = EditarUsuarioForm(instance=usuario)

    return render(request, 'editar_usuario.html', {'form': form, 'usuario': usuario})


@login_required
def excluir_usuario(request, usuario_id):
    """
    Exclui um Usuário do sistema.

    - Recupera o usuário com o ID fornecido.
    - Exclui o registro do usuário.
    - Exibe mensagem de sucesso ou erro.
    """
    usuario = get_object_or_404(Usuario, id=usuario_id)
    usuario.delete()
    messages.success(request, 'Usuário excluído com sucesso!')
    return redirect('listar_usuarios')


# ==================== ✅ VIEWS DE TUTOR ==================== #
@login_required
def cadastro_tutor(request):
    """
    Cadastra um novo tutor no sistema.

    - Processa o formulário de cadastro do tutor.
    - Valida os dados e verifica a existência de CPF duplicado.
    - Cria um novo registro de tutor.
    - Exibe mensagem de sucesso ou erro.
    """
    if request.method == 'POST':
        nome_completo = request.POST.get("nome")
        cpf = request.POST.get("cpf")
        telefone = request.POST.get("telefone")
        email = request.POST.get("email")
        endereco = request.POST.get("endereco")
        cidade = request.POST.get("cidade")
        estado = request.POST.get("estado")
        conheceu = request.POST.get("como_conheceu")

        if Tutor.objects.filter(cpf=cpf).exists():
            messages.error(request, 'Já existe um tutor cadastrado com esse CPF.')
            return redirect("cadastro_tutor")

        Tutor.objects.create(
            nome_completo=nome_completo,
            cpf=cpf,
            telefone=telefone,
            email=email,
            endereco=endereco,
            cidade=cidade,
            estado=estado,
            conheceu=conheceu,
        )
        messages.success(request, 'Tutor cadastrado com sucesso!')
        return redirect("cadastro_tutor")
    return render(request, 'cadastro_tutor.html')


@login_required
def editar_tutor(request, tutor_id):
    """
    Edita os dados de um tutor existente.

    - Recupera o tutor com o ID fornecido.
    - Processa o formulário de edição do tutor.
    - Valida os dados e atualiza o registro.
    - Exibe mensagem de sucesso ou erro.
    """
    tutor = get_object_or_404(Tutor, id=tutor_id)
    if request.method == 'POST':
        tutor.nome_completo = request.POST.get("nome")
        tutor.cpf = request.POST.get("cpf")
        tutor.telefone = request.POST.get("telefone")
        tutor.email = request.POST.get("email")
        tutor.endereco = request.POST.get("endereco")
        tutor.cidade = request.POST.get("cidade")
        tutor.estado = request.POST.get("estado")
        tutor.conheceu = request.POST.get("como_conheceu")
        tutor.observacoes = request.POST.get("observacoes")
        tutor.save()
        messages.success(request, 'Dados do tutor atualizados com sucesso!')
        return redirect('cadastro_tutor')
    return render(request, 'cadastro_tutor.html', {'tutor': tutor})


@login_required
def excluir_tutor(request, tutor_id):
    """
    Exclui um tutor do sistema.

    - Recupera o tutor com o ID fornecido.
    - Exclui o registro do tutor.
    - Exibe mensagem de sucesso ou erro.
    """
    tutor = get_object_or_404(Tutor, id=tutor_id)
    tutor.delete()
    messages.success(request, 'Tutor excluído com sucesso!')
    return redirect('listar_tutores')


@login_required
def listar_tutores(request):
    """
    Lista todos os tutores cadastrados no sistema, com opção de pesquisa.

    - Permite filtrar a listagem por nome completo ou CPF.
    - Exibe os tutores na página de listagem.
    """
    query = request.GET.get('q', '').strip()
    if query:
        tutores = Tutor.objects.filter(
            Q(nome_completo__icontains=query) | Q(cpf__icontains=query)
        )
    else:
        tutores = Tutor.objects.all()
    return render(request, 'listar_tutores.html', {'tutores': tutores})


@login_required
def detalhes_tutor(request, tutor_id):
    """
    Exibe os detalhes de um tutor específico e seus pets relacionados.

    - Recupera o tutor com o ID fornecido.
    - Recupera os pets associados ao tutor.
    - Exibe os detalhes na página de detalhes do tutor.
    """
    tutor = get_object_or_404(Tutor, id=tutor_id)
    pets = tutor.pet_set.all()
    return render(request, 'detalhes_tutor.html', {'tutor': tutor, 'pets': pets})


# ==================== ✅ VIEWS DE PET ==================== #
@login_required
def cadastro_pet(request):
    tutores = Tutor.objects.all()
    if request.method == 'POST':
        nome = request.POST.get("nome")
        especie = request.POST.get("especie")
        nova_especie = request.POST.get("nova_especie")
        raca = request.POST.get("raca")
        idade = request.POST.get("idade")
        sexo = request.POST.get("sexo")
        peso = request.POST.get("peso")
        observacoes_medicas = request.POST.get("observacoes_medicas")
        tutor_id = request.POST.get("tutor")
        foto = request.FILES.get("foto")

        # Substitui 'especie' pela nova espécie se for "Outro"
        if especie == "Outro" and nova_especie:
            especie = nova_especie.strip()

        if not tutor_id:
            messages.error(request, 'Selecione um tutor para o pet.')
            return redirect("cadastro_pet")

        if Pet.objects.filter(nome=nome, tutor_id=tutor_id).exists():
            messages.error(request, 'Já existe um pet com esse nome para o tutor selecionado.')
            return redirect("cadastro_pet")

        Pet.objects.create(
            nome=nome,
            especie=especie,
            raca=raca,
            idade=idade,
            sexo=sexo,
            peso=peso,
            observacoes_medicas=observacoes_medicas,
            tutor_id=tutor_id,
            foto=foto if foto else None,
        )
        messages.success(request, 'Pet cadastrado com sucesso!')
        return redirect("cadastro_pet")

    return render(request, 'cadastro_pet.html', {'tutores': tutores})


@login_required
def editar_pet(request, pet_id):
    """
    Edita os dados de um pet existente.

    - Recupera o pet com o ID fornecido.
    - Processa o formulário de edição do pet.
    - Valida os dados e atualiza o registro.
    - Exibe mensagem de sucesso ou erro.
    """
    pet = get_object_or_404(Pet, id=pet_id)
    tutores = Tutor.objects.all()
    if request.method == 'POST':
        pet.nome = request.POST.get("nome")
        especie = request.POST.get("especie")
        nova_especie = request.POST.get("nova_especie")
        
        # Substitui a espécie pelo valor digitado, se for "Outro"
        if especie == "Outro" and nova_especie:
            especie = nova_especie.strip()
        pet.especie = especie

        pet.raca = request.POST.get("raca")
        pet.idade = request.POST.get("idade")
        pet.sexo = request.POST.get("sexo")
        pet.peso = request.POST.get("peso") or None
        pet.observacoes_medicas = request.POST.get("observacoes_medicas")
        pet.tutor_id = request.POST.get("tutor")

        if 'foto' in request.FILES:
            pet.foto = request.FILES['foto']

        pet.save()
        messages.success(request, 'Pet atualizado com sucesso!')
        return redirect('cadastro_pet')

    return render(request, 'cadastro_pet.html', {'pet': pet, 'tutores': tutores})


@login_required
def excluir_pet(request, pet_id):
    """
    Exclui um pet do sistema.

    - Recupera o pet com o ID fornecido.
    - Exclui o registro do pet.
    - Exibe mensagem de sucesso ou erro.
    """
    pet = get_object_or_404(Pet, id=pet_id)
    pet.delete()
    messages.success(request, 'Pet excluído com sucesso!')
    return redirect('listar_pets')


@login_required
def listar_pets(request):
    """
    Lista todos os pets cadastrados no sistema, com opção de pesquisa.

    - Permite filtrar a listagem por nome do pet ou nome do tutor.
    - Exibe os pets na página de listagem.
    """
    query = request.GET.get('q', '')
    if query:
        pets = Pet.objects.filter(
            Q(nome__icontains=query) | Q(tutor__nome_completo__icontains=query)
        )
    else:
        pets = Pet.objects.all()
    return render(request, 'listar_pet.html', {'pets': pets})


# ==================== ✅ VIEWS DE SERVIÇO ==================== #
@login_required
def cadastro_servico(request):
    """
    Cadastra um novo serviço no sistema.

    - Processa o formulário de cadastro do serviço.
    - Valida os dados e verifica a existência de nome de serviço duplicado.
    - Cria um novo registro de serviço.
    - Exibe mensagem de sucesso ou erro.
    """
    if request.method == 'POST':
        nome_servico = request.POST.get("nome_servico")
        descricao = request.POST.get("descricao")
        preco = request.POST.get("preco")

        if Servico.objects.filter(nome_servico=nome_servico).exists():
            messages.error(request, 'Já existe um serviço cadastrado com esse nome.')
            return redirect("cadastro_servico")

        Servico.objects.create(
            nome_servico=nome_servico,
            descricao=descricao,
            preco=preco
        )
        messages.success(request, 'Serviço cadastrado com sucesso!')
        return redirect("cadastro_servico")
    return render(request, 'cadastro_servico.html')


@login_required
def editar_servicos(request, servicos_id):
    """
    Edita os dados de um serviço existente.

    - Recupera o serviço com o ID fornecido.
    - Processa o formulário de edição do serviço.
    - Valida os dados e atualiza o registro.
    - Exibe mensagem de sucesso ou erro.
    """
    servico = get_object_or_404(Servico, id=servicos_id)
    if request.method == 'POST':
        servico.nome_servico = request.POST.get("nome_servico")
        servico.descricao = request.POST.get("descricao")
        servico.preco = request.POST.get("preco")
        servico.save()
        messages.success(request, 'Serviço atualizado com sucesso!')
        return redirect('cadastro_servico')
    return render(request, 'cadastro_servico.html', {'servico': servico})


@login_required
def excluir_servicos(request, servicos_id):
    """
    Exclui um serviço do sistema.

    - Recupera o serviço com o ID fornecido.
    - Exclui o registro do serviço.
    - Exibe mensagem de sucesso ou erro.
    """
    servico = get_object_or_404(Servico, id=servicos_id)
    servico.delete()
    messages.success(request, 'Serviço excluído com sucesso!')
    return redirect('listar_servicos')


@login_required
def listar_servicos(request):
    """
    Lista todos os serviços cadastrados no sistema, com opção de pesquisa.

    - Permite filtrar a listagem por nome ou descrição do serviço.
    - Exibe os serviços na página de listagem.
    """
    query = request.GET.get('q', '')
    if query:
        servicos = Servico.objects.filter(
            Q(nome_servico__icontains=query) | Q(descricao__icontains=query)
        )
    else:
        servicos = Servico.objects.all()
    return render(request, 'listar_servicos.html', {'servicos': servicos, 'query': query})


# ==================== ✅ VIEWS DE AGENDAMENTO ==================== #
@login_required
def agendamento(request):
    """
    Cadastra um novo agendamento no sistema.

    - Processa o formulário de agendamento.
    - Valida os dados, incluindo a disponibilidade do pet.
    - Cria um novo registro de agendamento.
    - Calcula o valor total do agendamento.
    - Exibe mensagem de sucesso ou erro.
    """
    tutores = Tutor.objects.all()
    pets = Pet.objects.all()
    servicos = Servico.objects.all()

    if request.method == 'POST':
        tutor_id = request.POST.get('tutor_id')
        pet_id = request.POST.get('pet_id')
        data_agendamento = request.POST.get('data_agendamento')
        status = request.POST.get('status')
        observacoes = request.POST.get('observacoes')

        # Tratamento do desconto_percentual
        desconto_percentual = request.POST.get('desconto_percentual', 0) or 0
        try:
            desconto_percentual = float(desconto_percentual)
        except ValueError:
            desconto_percentual = 0

        servico_ids = request.POST.getlist('servico_id')
        servicos_selecionados = Servico.objects.filter(id__in=servico_ids)

        # Verificação: esse pet já está com outro tutor?
        try:
            pet = Pet.objects.get(id=pet_id)
            if pet.tutor_id != int(tutor_id):
                messages.error(request, f'O pet "{pet.nome}" já está cadastrado com outro tutor.')
                return render(request, 'agendamento.html', {
                    'tutores': tutores,
                    'pets': pets,
                    'servicos': servicos
                })
        except Pet.DoesNotExist:
            messages.error(request, 'Pet não encontrado.')
            return render(request, 'agendamento.html', {
                'tutores': tutores,
                'pets': pets,
                'servicos': servicos
            })

        # Criação do agendamento
        agendamento = Agendamento.objects.create(
            tutor_id=tutor_id,
            pet_id=pet_id,
            data_agendamento=data_agendamento,
            status=status,
            observacoes=observacoes,
            desconto_percentual=desconto_percentual,
        )

        # Adiciona os serviços ao agendamento
        agendamento.servicos.set(servicos_selecionados)

        # Calcula os valores
        valor_bruto = servicos_selecionados.aggregate(total=Sum('preco'))['total'] or 0
        agendamento.valor_total = agendamento.calcular_valor_total()
        agendamento.save()

        messages.success(request, 'Agendamento cadastrado com sucesso!')

        # Redireciona para limpar o formulário
        return redirect('agendamento')

    return render(request, 'agendamento.html', {
        'tutores': tutores,
        'pets': pets,
        'servicos': servicos
    })


@login_required
def editar_agendamentos(request, agendamentos_id):
    """
    Edita um agendamento existente.

    - Recupera o agendamento com o ID fornecido.
    - Processa o formulário de edição do agendamento.
    - Valida os dados e atualiza o registro.
    - Recalcula o valor total do agendamento.
    - Exibe mensagem de sucesso ou erro.
    """

    agendamento = get_object_or_404(Agendamento, id=agendamentos_id)
    tutores = Tutor.objects.all()
    pets = Pet.objects.all()
    servicos = Servico.objects.all()

    if request.method == 'POST':
        agendamento.tutor_id = request.POST.get("tutor_id")
        agendamento.pet_id = request.POST.get("pet_id")
        agendamento.data_agendamento = request.POST.get("data_agendamento")
        agendamento.status = request.POST.get("status")
        agendamento.observacoes = request.POST.get("observacoes")

        try:
            desconto_percentual = Decimal(request.POST.get('desconto_percentual', 0))
        except InvalidOperation:
            messages.error(request, 'Por favor, insira um valor de desconto válido.')
            return render(request, 'editar_agendamento.html', {
                'agendamento': agendamento,
                'tutores': tutores,
                'pets': pets,
                'servicos': servicos
            })

        servico_ids = request.POST.getlist("servico_id")
        servicos_selecionados = Servico.objects.filter(id__in=servico_ids)

        agendamento.servicos.set(servicos_selecionados)
        agendamento.desconto_percentual = desconto_percentual
        agendamento.valor_total = agendamento.calcular_valor_total()
        agendamento.save()
        messages.success(request, 'Agendamento atualizado com sucesso!')
        return redirect('agendamento')

    return render(request, 'editar_agendamento.html', {
        'agendamento': agendamento,
        'tutores': tutores,
        'pets': pets,
        'servicos': servicos
    })


@login_required
def excluir_agendamentos(request, agendamentos_id):
    """
    Exclui um agendamento do sistema.

    - Recupera o agendamento com o ID fornecido.
    - Exclui o registro do agendamento.
    - Exibe mensagem de sucesso ou erro.
    """
    agendamento = get_object_or_404(Agendamento, id=agendamentos_id)
    agendamento.delete()
    messages.success(request, 'Agendamento excluído com sucesso!')
    return redirect('listar_agendamentos')


@login_required
def listar_agendamentos(request):
    """
    Lista todos os agendamentos cadastrados no sistema, com opção de pesquisa.

    - Permite filtrar a listagem por nome do tutor, nome do pet ou nome do serviço.
    - Exibe os agendamentos na página de listagem.
    """
    query = request.GET.get('q', '')
    if query:
        agendamentos = Agendamento.objects.filter(
            Q(tutor__nome_completo__icontains=query) |
            Q(pet__nome__icontains=query) |
            Q(servicos__nome_servico__icontains=query)
        ).distinct().prefetch_related('servicos', 'tutor', 'pet')
    else:
        agendamentos = Agendamento.objects.all().prefetch_related('servicos', 'tutor', 'pet')
    return render(request, 'listar_agendamentos.html', {'agendamentos': agendamentos})


# ==================== ✅ VIEWS DE RELATÓRIOS ==================== #
@login_required
def relatorios(request):
    # Verifica se o usuário está no grupo "Administrador"
    if request.user.nivel != 'admin':
        messages.error(request, "Você não tem permissão para acessar esta funcionalidade.")
        return redirect('home')

    tutores = Tutor.objects.all()
    agendamentos = Agendamento.objects.select_related('tutor', 'pet').prefetch_related('servicos')

    tutor_id = request.GET.get('tutor_id')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if tutor_id:
        agendamentos = agendamentos.filter(tutor_id=tutor_id)
    if data_inicio:
        agendamentos = agendamentos.filter(data_agendamento__gte=data_inicio)
    if data_fim:
        agendamentos = agendamentos.filter(data_agendamento__lte=data_fim)

    total_geral = agendamentos.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal(0)

    context = {
        'tutores': tutores,
        'agendamentos': agendamentos,
        'total_geral': total_geral,
    }
    return render(request, 'relatorios.html', context)


@login_required
def exportar_relatorio_excel(request):
    """
    Gera e exporta relatórios de agendamentos em formato Excel.
    - Aplica os mesmos filtros da view 'relatorios'.
    - Cria uma planilha com os dados dos agendamentos.
    - Formata a planilha e prepara a resposta HTTP para download.
    """
    # Verificação de permissão
    if request.user.nivel != 'admin':
        messages.error(request, "Você não tem permissão para exportar relatórios.")
        return redirect('home')

    agendamentos = Agendamento.objects.select_related('tutor', 'pet').prefetch_related('servicos')

    tutor_id = request.GET.get('tutor_id')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if tutor_id:
        agendamentos = agendamentos.filter(tutor_id=tutor_id)
    if data_inicio:
        agendamentos = agendamentos.filter(data_agendamento__gte=data_inicio)
    if data_fim:
        agendamentos = agendamentos.filter(data_agendamento__lte=data_fim)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Agendamentos"

    headers = ["Data", "Tutor", "Pet", "Serviços", "Desconto %", "Total (R$)"]
    ws.append(headers)

    for agendamento in agendamentos:
        servicos_str = ", ".join(
            f"{s.nome_servico} (R$ {s.preco})" for s in agendamento.servicos.all()
        )
        ws.append([
            agendamento.data_agendamento.strftime('%d/%m/%Y'),
            agendamento.tutor.nome_completo,
            agendamento.pet.nome,
            servicos_str,
            agendamento.desconto_percentual,
            float(agendamento.valor_total)
        ])

    for column_cells in ws.columns:
        max_length = 0
        col = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col].width = max_length + 3

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_agendamentos.xlsx'
    wb.save(response)
    return response


# ==================== ✅ VIEWS DE NOTA DE SERVIÇO ==================== #
@login_required
def nota_atendimento(request, agendamento_id):
    """
    Exibe a nota de atendimento de um agendamento.
    - Recupera o agendamento com os dados relacionados (serviços, pet e tutor).
    - Calcula o valor bruto, desconto e valor total com desconto.
    - Exibe a nota de atendimento em HTML.
    """
    agendamento = get_object_or_404(
        Agendamento.objects.select_related('pet__tutor').prefetch_related('servicos'),
        id=agendamento_id
    )

    valor_bruto = sum(Decimal(servico.preco) for servico in agendamento.servicos.all())
    desconto_percentual = agendamento.desconto_percentual
    valor_desconto = (valor_bruto * desconto_percentual / 100) if desconto_percentual else Decimal(0)
    valor_total_com_desconto = valor_bruto - valor_desconto

    return render(request, 'nota_atendimento.html', {
        'agendamento': agendamento,
        'valor_bruto': valor_bruto,
        'valor_desconto': valor_desconto,
        'valor_total_com_desconto': valor_total_com_desconto
    })
    

@login_required
def gerar_nota(request, agendamento_id):
    """
    Gera e exporta a nota de atendimento em formato PDF.
    - Recupera o agendamento com os dados relacionados.
    - Calcula o valor total e o desconto.
    - Gera o HTML para o PDF usando um template.
    - Converte o HTML para PDF usando pdfkit.
    - Prepara a resposta HTTP para download do PDF.
    """
    agendamento = get_object_or_404(
        Agendamento.objects.select_related('pet__tutor').prefetch_related('servicos'),
        id=agendamento_id
    )

    valor_total = sum(servico.preco for servico in agendamento.servicos.all())
    desconto_percentual = agendamento.desconto_percentual
    valor_desconto = (valor_total * desconto_percentual) / 100 if desconto_percentual else 0
    valor_total_com_desconto = valor_total - valor_desconto

    html_string = render_to_string('nota_pdf.html', {
        'agendamento': agendamento,
        'valor_total': valor_total,
        'valor_desconto': valor_desconto,
        'valor_total_com_desconto': valor_total_com_desconto,
        'logo_path': os.path.join(settings.BASE_DIR, 'static/img/klug_pet_logo.png')
    })

    pdf = pdfkit.from_string(html_string, False)

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="nota_{agendamento.id}.pdf"'
    return response
